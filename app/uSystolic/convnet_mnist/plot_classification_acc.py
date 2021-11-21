from openpyxl import load_workbook
import numpy as np
import os.path
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager

class bcolors:
    """
    Reference from: 
    https://svn.blender.org/svnroot/bf-blender/trunk/blender/build_files/scons/tools/bcolors.py
    """
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def get_lists(tab=None, cfg="tlut"):
    """
    Returns: 
        loss_list, acc_list
    """
    if cfg == 'tlut':
        # loss = get_cell_vals(tab, 'B6:B20')
        loss = [0.0296, 0.044, 0.0612, 0.0997, 0.1713, 0.2981, 1.6251, 1.9081, 2.0437, 1.8245, 2.1162, 2.2531, 2.2789, 2.3028, 2.3035]
        # acc = get_cell_vals(tab, 'C6:C20')
        acc = [99.04, 99.04, 99, 98.94, 98.77, 98.58, 83.61, 72.43, 58.83, 81.77, 60.81, 20.46, 12.88, 11.35, 11.35]
        return loss, acc
    elif cfg == "fxp": 
        # loss = get_cell_vals(tab, 'B3')
        # loss = np.append([np.nan]*14, [0.0296])
        loss = [0.0296]*15
        # acc = get_cell_vals(tab, 'C3')
        # acc = np.append([np.nan]*14, [99.04])
        acc = [99.04]*15
        return loss, acc
    elif cfg == 'fp':
        # loss = get_cell_vals(tab, 'B2')
        # loss = np.append([np.nan]*14, [0.027])
        loss = [0.027]*15
        # acc = get_cell_vals(tab, 'C2')
        # acc = np.append([np.nan]*14, [99.11])
        acc = [99.11]*15
        
        return loss, acc
    elif cfg == 'hub':
        # loss = get_cell_vals(tab, 'B4')
        # loss = np.append([np.nan]*14, [0.0306])
        loss = [0.0306]*15
        # acc = get_cell_vals(tab, 'C4')
        # acc = np.append([np.nan]*14, [98.98])
        acc = [98.98]*15
        return loss, acc
    else:
        print(bcolors.FAIL + f'Unrecognized cfg {cfg}. Options: tlut, fxp, fp, hub' + bcolors.ENDC)
        exit()
    
def get_cell_vals(tab, str_range):
    if ':' in str_range:
        cells = tab[str_range]
        return [x.value for cell in cells for x in cell]
    else:
        cell = tab[str_range]
        return cell.value

def query_workbook(workbook=None, str_tab=None, cfg='ubrain'):
    sheets = workbook.sheetnames
    ind = sheets.index(str_tab)
    return get_lists(tab=workbook[sheets[ind]], cfg=cfg)

def plot_design(design='tlut', absolute_path=None, filename=None):
    # workbook = load_workbook(filename=absolute_path+filename, data_only=True)
    # loss_, acc_ = query_workbook(workbook, 'Sheet1', design)
    # return loss_, acc_
    return get_lists(None, design)

def plot_fig(list_,str_in):
    """    
    [fp, fxp, hub, tlut]
    """
    font = {'family':'Times New Roman', 'size': 6}
    matplotlib.rc('font', **font)
    x_axis = range(1, 16)
    my_dpi = 300
    fig_h = 1.1
    fig_w = 3.3115

    # TODO: Change color
    tlut_color = "#6ACCBC"
    fp_color = "#FF7F7F"
    fxp_color = "#D783FF"
    hub_color = '#7A81FF'

    width = 0.20
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    # x = [15,14,13,12,11,10,9,8,7,6,5,4,3,2,1]
    x = range(15, 0, -1)
    # print(x); exit()

    ax.plot(x, list_[0], "-o", label="FP", color=fp_color, ms=4)

    ax.plot(x, list_[1], "-o", label="FXP", color=fxp_color, ms=4)

    ax.plot(x, list_[2], "-s", label="uSystolic", color=hub_color, ms=4)

    ax.plot(x, list_[3], "-^", label="Temporal-LUT", color=tlut_color, ms=4)

    ax.set_xticks(x)
    # ax.set_xticklabels(x_label)
    ax.set_yscale('linear')
    if str_in == "loss":
        ax.set_yticks([0, 0.02, 0.04, 0.06, 0.08])
        ax.set_yticklabels(["0.00", "0.02", "0.04", "0.06", "0.08"])
        ax.set_ylim(0.02, 0.1)
        
    else:
        ax.legend(ncol=4, frameon=False)
        ax.set_yticks([85, 90, 95, 100])
        ax.set_yticklabels(["85", "90", "95", "100"])
        ax.set_ylim(85, 103)

    fig.tight_layout()
    plt.show()
    path = "/home/zhewen/Repo/UnarySim/app/uSystolic/convnet_mnist/"
    fig.savefig(path + f"{str_in}.pdf", bbox_inches='tight', dpi=my_dpi, pad_inches=0.02)

def main():
    # NOTE: Update absolute path and filename
    path = '/home/zhewen/Repo/UnarySim/app/uSystolic/convnet_mnist/'
    filename = "accuracy.xlsx"
    file_exists = os.path.exists(path+filename)
    if file_exists == False:
        print(bcolors.FAIL + f'{filename} does not exist at path {path}. Did you forget to put it in?' + bcolors.ENDC)
        exit()
    else: print(bcolors.OKGREEN + f'Processing {path+filename}...' + bcolors.ENDC)

    loss_hub, acc_hub = plot_design('hub', path, filename)
    loss_tlut, acc_tlut = plot_design('tlut', path, filename)
    loss_fxp, acc_fxp = plot_design('fxp', path, filename)
    loss_fp, acc_fp = plot_design('fp', path, filename)
    plot_fig([loss_fp, loss_fxp, loss_hub, loss_tlut], 'loss')
    plot_fig([acc_fp, acc_fxp, acc_hub, acc_tlut], 'accuracy')

if __name__ == "__main__":
    main()
